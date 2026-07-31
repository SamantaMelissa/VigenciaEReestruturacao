from collections import defaultdict
from datetime import datetime
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "Dados" / "T.I. Dados dos cursos 2026.xlsx"
OUTPUT = ROOT / "assets" / "data" / "courses-data.js"
ROOT_DECISIONS_SOURCE = ROOT / "Definição da Situação dos Cursos.xlsx"
DECISIONS_SOURCE = (
    ROOT_DECISIONS_SOURCE
    if ROOT_DECISIONS_SOURCE.exists()
    else ROOT / "Dados" / "Definição da Situação dos Cursos.xlsx"
)
DECISIONS_SEED_OUTPUT = ROOT / "supabase" / "seed_initial_decisions.sql"
DECISIONS_FIX_OUTPUT = ROOT / "supabase" / "fix_initial_decision_criteria.sql"
COURSE_SCOPE_OUTPUT = ROOT / "supabase" / "sync_course_analysis_scope.sql"
OUT_OF_GRID_SOURCE = ROOT / "config" / "courses-out-of-grid.txt"


def load_code_list(path):
    if not path.exists():
        return set()
    return {
        line.strip()
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    }


OUT_OF_GRID_CODES = load_code_list(OUT_OF_GRID_SOURCE)

wb = load_workbook(SOURCE, data_only=True, read_only=True)
export = wb["Export"]
headers = [cell.value for cell in next(export.iter_rows())]
header_index = {str(value).strip(): idx for idx, value in enumerate(headers) if value}

enrollments = defaultdict(lambda: {2023: 0, 2024: 0, 2025: 0, "units": set()})
mat = wb["Matrículas"]
mat_headers = [cell.value for cell in next(mat.iter_rows())]
mat_index = {str(value).strip(): idx for idx, value in enumerate(mat_headers) if value}

for row in mat.iter_rows(min_row=2, values_only=True):
    code = row[mat_index["Código do curso"]]
    if code in (None, ""):
        continue
    key = str(int(code)) if isinstance(code, (int, float)) else str(code).strip()
    for year in (2023, 2024, 2025):
        value = row[mat_index[str(year)]] or 0
        if isinstance(value, (int, float)):
            enrollments[key][year] += int(value)
    unit = row[mat_index["Rótulos de Linha"]]
    if unit not in (None, ""):
        enrollments[key]["units"].add(str(unit))


def value(row, name):
    idx = header_index.get(name)
    return row[idx] if idx is not None and idx < len(row) else None


def date_string(raw):
    if isinstance(raw, datetime):
        return raw.strftime("%d/%m/%Y")
    return str(raw) if raw not in (None, "") else ""


def criterion(course_type, level):
    text = f"{course_type or ''} {level or ''}".lower()
    fic_specific = ("aperfeiçoamento" in text or "iniciação" in text or
                    "especialização profissional" in text)
    return "fic" if fic_specific else "regular"


def normalized_result(raw):
    result = str(raw).strip().upper()
    corrections = {
        "RESTRUTURAR": "REESTRUTURAR",
        "FECHAR VIGÊNCIA": "FECHAR A VIGÊNCIA",
    }
    return corrections.get(result, result)


courses = []
seen = set()
for row in export.iter_rows(min_row=2, values_only=True):
    name = value(row, "Curso")
    code = value(row, "Código do Curso")
    if not name or code in (None, ""):
        continue
    code_key = str(int(code)) if isinstance(code, (int, float)) else str(code).strip()
    unique = (code_key, str(name).strip().upper())
    if unique in seen:
        continue
    seen.add(unique)
    enrollment = enrollments[code_key]
    course_type = value(row, "Tipo de Curso")
    level = value(row, "Nível")
    courses.append({
        "code": code_key,
        "name": str(name).strip().upper(),
        "hours": value(row, "C. H.") or "",
        "level": level or "",
        "strategy": value(row, "Estratégia") or "",
        "type": course_type or "",
        "area": value(row, "Área") or "",
        "segment": value(row, "Segmento de Área") or "",
        "creator": value(row, "Unidade criadora") or "",
        "start": date_string(value(row, "Início de vigęncia") or value(row, "Início de vigência")),
        "end": date_string(value(row, "Data de término")),
        "criterion": criterion(course_type, level),
        "enrollments": {
            "2023": enrollment[2023],
            "2024": enrollment[2024],
            "2025": enrollment[2025],
        },
        "units": len(enrollment["units"]),
        "unitCodes": sorted(
            enrollment["units"],
            key=lambda item: (not item.isdigit(), int(item) if item.isdigit() else item)
        ),
    })

courses.sort(key=lambda item: item["name"])
with open(OUTPUT, "w", encoding="utf-8", newline="\n") as handle:
    handle.write("window.COURSES_DATA = ")
    json.dump(courses, handle, ensure_ascii=False, separators=(",", ":"))
    handle.write(";\n")

print(f"{len(courses)} cursos gravados em {OUTPUT}")

course_scope = [
    {
        "code": item["code"],
        "name": item["name"],
        "creatorUnit": str(item["creator"]).strip(),
        "isAnalyzable": (
            str(item["creator"]).strip().upper() == "GED"
            and item["code"] not in OUT_OF_GRID_CODES
        ),
    }
    for item in courses
]
course_scope_json = json.dumps(course_scope, ensure_ascii=False, separators=(",", ":"))
course_scope_sql = f"""-- Sincroniza o escopo de cursos que podem ser analisados.
-- Regra vigente: Unidade criadora = GED, exceto cursos retirados da grade.
-- O script não apaga dados e pode ser executado novamente com segurança.

begin;

create table if not exists public.course_analysis_scope (
  course_code text primary key,
  course_name text not null,
  creator_unit text not null,
  is_analyzable boolean not null,
  updated_at timestamptz not null default now()
);

alter table public.course_analysis_scope enable row level security;

drop policy if exists "Authenticated users view course analysis scope"
on public.course_analysis_scope;
create policy "Authenticated users view course analysis scope"
on public.course_analysis_scope for select to authenticated
using (true);

drop policy if exists "Managers maintain course analysis scope"
on public.course_analysis_scope;
create policy "Managers maintain course analysis scope"
on public.course_analysis_scope for all to authenticated
using (public.current_user_role() in ('gestor', 'admin'))
with check (public.current_user_role() in ('gestor', 'admin'));

grant select, insert, update, delete on public.course_analysis_scope to authenticated;

with source as (
  select *
  from jsonb_to_recordset($courses$
{course_scope_json}
$courses$::jsonb) as item(
    code text,
    name text,
    "creatorUnit" text,
    "isAnalyzable" boolean
  )
)
insert into public.course_analysis_scope (
  course_code,
  course_name,
  creator_unit,
  is_analyzable,
  updated_at
)
select
  code,
  name,
  "creatorUnit",
  "isAnalyzable",
  now()
from source
on conflict (course_code) do update set
  course_name = excluded.course_name,
  creator_unit = excluded.creator_unit,
  is_analyzable = excluded.is_analyzable,
  updated_at = now();

commit;

select
  count(*) as total_cursos,
  count(*) filter (where is_analyzable) as cursos_ged_para_analise,
  count(*) filter (where not is_analyzable) as cursos_fora_da_analise
from public.course_analysis_scope;
"""

with open(COURSE_SCOPE_OUTPUT, "w", encoding="utf-8", newline="\n") as handle:
    handle.write(course_scope_sql)

print(f"script de escopo GED gravado em {COURSE_SCOPE_OUTPUT}")

decision_wb = load_workbook(DECISIONS_SOURCE, data_only=True, read_only=True)
decision_sheet = decision_wb.active
decision_rows = decision_sheet.iter_rows(values_only=True)
decision_headers = [cell for cell in next(decision_rows)]
decision_index = {str(value).strip(): idx for idx, value in enumerate(decision_headers) if value}
decisions = []
contacts = []

for order, row in enumerate(decision_rows, start=1):
    code = row[decision_index["Código do Curso"]]
    name = row[decision_index["Curso"]]
    result = row[decision_index["Situação do Curso"]]
    if code in (None, "") or not name:
        continue
    code_key = str(int(code)) if isinstance(code, (int, float)) else str(code).strip()
    justification = row[decision_index["Justificativa"]] or ""
    observations = row[decision_index["OBSERVAÇÕES"]] or ""
    decision_type = row[decision_index["Tipo de Curso"]]
    decision_level = row[decision_index["Nível"]]
    decision_criterion = criterion(decision_type, decision_level)
    contact_text = f"{justification} {observations}".upper()
    if "CONVERSAR COM O COORDENADOR" in contact_text:
        enrollment = enrollments[code_key]
        contacts.append({
            "sourceId": f"definicao-contato-{code_key}",
            "code": code_key,
            "name": str(name).strip().upper(),
            "criterionKey": decision_criterion,
            "criterion": (
                "Critério FIC — Aperfeiçoamento, Especialização e Iniciação"
                if decision_criterion == "fic"
                else "Critério Regular / Qualificação"
            ),
            "units": sorted(enrollment["units"]),
            "enrollments": {
                "2023": enrollment[2023],
                "2024": enrollment[2024],
                "2025": enrollment[2025],
            },
            "reason": str(justification).strip(),
            "observations": str(observations).strip(),
        })
    if not result:
        continue
    decisions.append({
        "sourceId": f"definicao-{code_key}",
        "id": 800000000 + order,
        "date": "Planilha de definição",
        "code": code_key,
        "name": str(name).strip().upper(),
        "criterionKey": decision_criterion,
        "criterion": (
            "Critério FIC — Aperfeiçoamento, Especialização e Iniciação"
            if decision_criterion == "fic"
            else "Critério Regular / Qualificação"
        ),
        "result": normalized_result(result),
        "justification": str(justification).strip(),
        "observations": str(observations).strip(),
        "source": "Definição da Situação dos Cursos.xlsx",
    })

seed_json = json.dumps(decisions, ensure_ascii=False, separators=(",", ":"))
seed_relation = f"""select *
from jsonb_to_recordset($seed$
{seed_json}
$seed$::jsonb) as item(
  "sourceId" text,
  id bigint,
  date text,
  code text,
  name text,
  "criterionKey" text,
  criterion text,
  result text,
  justification text,
  observations text,
  source text
)"""
contact_json = json.dumps(contacts, ensure_ascii=False, separators=(",", ":"))
contact_relation = f"""select *
from jsonb_to_recordset($contacts$
{contact_json}
$contacts$::jsonb) as item(
  "sourceId" text,
  code text,
  name text,
  "criterionKey" text,
  criterion text,
  units jsonb,
  enrollments jsonb,
  reason text,
  observations text
)"""
seed_sql = f"""-- Carga inicial das decisões existentes
-- Execute depois de schema.sql e da criação do primeiro administrador.
-- O script é idempotente: sourceId impede duplicações.

begin;

do $$
begin
  if not exists (select 1 from public.profiles where role = 'admin') then
    raise exception 'Crie o primeiro administrador antes de executar esta carga.';
  end if;
end;
$$;

with seed as (
{seed_relation}
)
update public.evaluations evaluation
set
  course_code = seed.code,
  course_name = seed.name,
  criterion_key = seed."criterionKey",
  criterion_label = seed.criterion,
  final_result = seed.result,
  justification = seed.justification,
  state = coalesce(evaluation.state, '{{}}'::jsonb) || jsonb_build_object(
    'sourceId', seed."sourceId",
    'source', seed.source,
    'sourceDate', seed.date,
    'observations', seed.observations,
    'imported', true
  ),
  updated_at = now()
from seed
where evaluation.state ->> 'sourceId' = seed."sourceId";

with seed as (
{seed_relation}
)
insert into public.evaluations (
  course_code,
  course_name,
  criterion_key,
  criterion_label,
  status,
  current_question,
  final_result,
  justification,
  state,
  created_by,
  created_at,
  updated_at,
  completed_at
)
select
  seed.code,
  seed.name,
  seed."criterionKey",
  seed.criterion,
  'concluida',
  null,
  seed.result,
  seed.justification,
  jsonb_build_object(
    'sourceId', seed."sourceId",
    'source', seed.source,
    'sourceDate', seed.date,
    'observations', seed.observations,
    'imported', true
  ),
  (select id from public.profiles where role = 'admin' order by created_at limit 1),
  now(),
  now(),
  now()
from seed
where not exists (
  select 1
  from public.evaluations existing
  where existing.state ->> 'sourceId' = seed."sourceId"
     or (existing.course_code = seed.code and existing.status = 'concluida')
);

-- Inclui somente os contatos ainda ausentes. Contatos existentes, inclusive os
-- já concluídos, permanecem intactos.
with contacts as (
{contact_relation}
)
insert into public.school_validations (
  evaluation_id,
  course_code,
  course_name,
  criterion_key,
  criterion_label,
  units,
  enrollments,
  reason_question,
  decision_trail,
  status,
  notes,
  created_by,
  created_at,
  updated_at
)
select
  (
    select evaluation.id
    from public.evaluations evaluation
    where evaluation.course_code = contacts.code
    order by evaluation.updated_at desc
    limit 1
  ),
  contacts.code,
  contacts.name,
  contacts."criterionKey",
  contacts.criterion,
  array(select jsonb_array_elements_text(contacts.units)),
  contacts.enrollments,
  'A escola tem uma justificativa técnica para manter o curso?',
  jsonb_build_array(jsonb_build_object(
    'sourceId', contacts."sourceId",
    'source', 'Definição da Situação dos Cursos.xlsx',
    'reason', contacts.reason
  )),
  'pendente',
  nullif(contacts.observations, ''),
  (select id from public.profiles where role = 'admin' order by created_at limit 1),
  now(),
  now()
from contacts
where not exists (
  select 1
  from public.school_validations existing
  where existing.course_code = contacts.code
);

commit;

-- Conferência das {len(decisions)} decisões preenchidas na planilha. O total
-- considera tanto importações quanto análises já concluídas no sistema.
with expected(code) as (
  select value #>> '{{}}'
  from jsonb_array_elements('{json.dumps([item["code"] for item in decisions])}'::jsonb)
)
select
  count(*) filter (where exists (
    select 1
    from public.evaluations evaluation
    where evaluation.course_code = expected.code
      and evaluation.status = 'concluida'
  )) as cursos_com_decisao_concluida,
  {len(decisions)} as decisoes_esperadas
from expected;

-- Conferência: devem existir os {len(contacts)} cursos da planilha na central,
-- contando os que já estavam cadastrados antes desta sincronização.
with expected(code) as (
  select value #>> '{{}}'
  from jsonb_array_elements('{json.dumps([item["code"] for item in contacts])}'::jsonb)
)
select
  count(*) as contatos_da_planilha_no_banco,
  {len(contacts)} as contatos_esperados
from public.school_validations validation
where validation.course_code in (select code from expected);
"""

with open(DECISIONS_SEED_OUTPUT, "w", encoding="utf-8", newline="\n") as handle:
    handle.write(seed_sql)

print(f"script de carga gravado em {DECISIONS_SEED_OUTPUT}")

criteria_fix_data = [
    {
        "sourceId": item["sourceId"],
        "criterionKey": item["criterionKey"],
        "criterion": item["criterion"],
    }
    for item in decisions
]
criteria_fix_json = json.dumps(criteria_fix_data, ensure_ascii=False, separators=(",", ":"))
criteria_fix_sql = f"""-- Corrige o critério das decisões importadas anteriormente.
-- Pode ser executado com segurança mais de uma vez.

with correction as (
  select *
  from jsonb_to_recordset($fix$
{criteria_fix_json}
$fix$::jsonb) as item(
    "sourceId" text,
    "criterionKey" text,
    criterion text
  )
)
update public.evaluations evaluation
set
  criterion_key = correction."criterionKey",
  criterion_label = correction.criterion,
  updated_at = now()
from correction
where evaluation.state ->> 'sourceId' = correction."sourceId";

select criterion_key, criterion_label, count(*) as quantidade
from public.evaluations
where state ->> 'imported' = 'true'
group by criterion_key, criterion_label
order by criterion_key;
"""

with open(DECISIONS_FIX_OUTPUT, "w", encoding="utf-8", newline="\n") as handle:
    handle.write(criteria_fix_sql)

print(f"script de correção gravado em {DECISIONS_FIX_OUTPUT}")
